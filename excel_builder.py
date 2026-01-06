from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from typing import Any


def style_header(row):
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", fill_type="solid")


def autosize(ws):
    for col in ws.columns:
        try:
            length = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = length + 2
        except Exception:
            continue


def _is_list_of_dicts(obj):
    return isinstance(obj, list) and (len(obj) == 0 or isinstance(obj[0], dict))


def build_excel(data: Any, filename="daily_report.xlsx"):
    """
    Accepts either:
      - A dict mapping sheet_name -> list_of_dicts
      - OR the old signature: (users, posts, todos) passed as positional args (for backward compat)

    If data is a dict:
      - Creates one sheet per key
      - Uses keys of first dict as headers
      - Each dict becomes one row
    """

    wb = Workbook()

    if isinstance(data, dict):
        first_sheet = True

        for sheet_name, dataset in data.items():
            # Only create sheets for list-of-dict datasets
            if not _is_list_of_dicts(dataset):
                continue

            if first_sheet:
                ws = wb.active
                ws.title = str(sheet_name)[:31]  # Excel sheet name max length
                first_sheet = False
            else:
                ws = wb.create_sheet(str(sheet_name)[:31])

            # If dataset is empty
            if not dataset:
                ws.append(["No data"])
                continue

            # Headers from first row keys
            headers = list(dataset[0].keys())
            ws.append(headers)
            style_header(ws[1])

            for item in dataset:
                row = [item.get(h, "") for h in headers]
                ws.append(row)

            autosize(ws)

    else:
        # Backward compatibility mode (users, posts, todos)
        try:
            users, posts, todos = data

            # Users sheet
            ws = wb.active
            ws.title = "Users"
            ws.append(["ID", "Name", "Username", "Email"])
            style_header(ws[1])
            for u in users:
                ws.append([u["id"], u["name"], u["username"], u["email"]])
            autosize(ws)

            # Posts sheet
            ws2 = wb.create_sheet("Posts")
            ws2.append(["ID", "UserID", "Title"])
            style_header(ws2[1])
            for p in posts:
                ws2.append([p["id"], p["userId"], p["title"]])
            autosize(ws2)

            # Todos sheet
            ws3 = wb.create_sheet("Todos")
            ws3.append(["ID", "UserID", "Title", "Completed"])
            style_header(ws3[1])
            for t in todos:
                ws3.append([t["id"], t["userId"], t["title"], t["completed"]])
            autosize(ws3)

        except Exception:
            raise ValueError("Unsupported data format for build_excel")

    wb.save(filename)
    return filename
